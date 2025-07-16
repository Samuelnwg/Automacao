
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl

# === CONFIGURAÇÕES ===
CAMINHO_PLANILHA = "Samuel C..xlsx"
ABA = "GOVSP"
COLUNA_CPF = "CPF "
COLUNA_OBSERVACAO = "Observação"
URL_PORTAL = "https://www.portaldoconsignado.org.br/home?40"

# === ABRE NAVEGADOR ===
options = Options()
options.add_experimental_option("detach", True)  # Mantém navegador aberto após execução
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# === ABRE SITE E ESPERA LOGIN MANUAL ===
driver.get(URL_PORTAL)
print("➡️ Faça login manualmente e clique em 'Consulta de Margem' duas vezes.")
input("✅ Pressione ENTER aqui depois de estar na tela de consulta de CPF...")

# === CARREGA PLANILHA ===
wb = openpyxl.load_workbook(CAMINHO_PLANILHA)
ws = wb[ABA]

# Pega índice das colunas
header = [cell.value for cell in ws[1]]
idx_cpf = header.index(COLUNA_CPF) + 1
idx_obs = header.index(COLUNA_OBSERVACAO) + 1

# === LOOP SOBRE LINHAS ===
for row in range(2, ws.max_row + 1):
    cpf = ws.cell(row=row, column=idx_cpf).value
    if not cpf:
        continue

    try:
        # Localiza o campo CPF e pesquisa
        campo_cpf = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "cpfServidor"))
        )
        campo_cpf.clear()
        campo_cpf.send_keys(str(cpf))
        botao_pesquisar = driver.find_element(By.NAME, "botaoPesquisar")
        botao_pesquisar.click()

        time.sleep(3)

        # Localiza a tabela de margens
        tabela = driver.find_element(By.XPATH, "//div[@id='painelMargensDisponiveis']")
        linhas = tabela.find_elements(By.TAG_NAME, "tr")

        valores = {}
        for linha in linhas[1:]:
            colunas = linha.find_elements(By.TAG_NAME, "td")
            nome_produto = colunas[0].text.strip()
            valor_str = colunas[1].text.strip().replace(".", "").replace(",", ".")
            try:
                valor = float(valor_str)
                valores[nome_produto] = valor
            except:
                continue

        fac = valores.get("CONSIGNACOES FACULTATIVAS", 0)
        cc = valores.get("CARTAO DE CREDITO", 0)
        cb = valores.get("CARTÃO DE BENEFÍCIO", 0)

        if fac >= 30:
            resultado = f"{fac:.2f}".replace(".", ",")
        elif cc >= 30 or cb >= 30:
            maior = max(cc, cb)
            resultado = f"{maior:.2f}".replace(".", ",", " ", maior)
        else:
            resultado = "Sem margem"

        ws.cell(row=row, column=idx_obs, value=resultado)
        print(f"✅ CPF {cpf}: {resultado}")

    except Exception as e:
        ws.cell(row=row, column=idx_obs, value="Erro")
        print(f"❌ Erro ao processar CPF {cpf}: {e}")

# === SALVA NOVO ARQUIVO ===
wb.save("Samuel_C_RESULTADO.xlsx")
print("✅ FIM! Resultados salvos em Samuel_C_RESULTADO.xlsx")
